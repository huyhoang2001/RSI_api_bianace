import os
import json
import concurrent.futures
import pandas as pd
import numpy as np
import ta
import gspread
import requests
from binance.client import Client
from dotenv import load_dotenv
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from oauth2client.service_account import ServiceAccountCredentials
from enum import Enum

load_dotenv()

ALLOWED_INTERVALS = ["5m", "15m", "30m", "1h", "4h", "1d", "1w"]


class DivergencePhase(Enum):
    IDLE = "IDLE"
    FORMING = "FORMING"
    DEVELOPING = "DEVELOPING"
    CONFIRMED = "CONFIRMED"


def ask_rsi_period():
    while True:
        user_input = input("Nhập RSI period (nhập 00 để thoát): ").strip()
        if user_input == "00":
            return None
        try:
            period = int(user_input)
            if period <= 0:
                raise ValueError
            return period
        except ValueError:
            print("Vui lòng nhập số nguyên dương.")


def ask_intervals():
    while True:
        print(f"\nKhung thời gian cho phép: {' '.join(ALLOWED_INTERVALS)}")
        user_input = input("Nhập khung thời gian (00 để quay lại): ").strip()
        if user_input == "00":
            return None
        intervals = user_input.split()
        if all(i in ALLOWED_INTERVALS for i in intervals) and intervals:
            return intervals
        print("Khung thời gian không hợp lệ.")


def ask_analysis_mode():
    while True:
        print("\n" + "="*50)
        print("1. RSI cơ bản (RSI >= 80 hoặc <= 20)")
        print("2. RSI Divergence (Phân kỳ theo rules V4)")
        print("3. Cả hai")
        print("00. Thoát")
        print("="*50)
        choice = input("Chọn chế độ: ").strip()
        if choice == "00":
            return None
        if choice in ["1", "2", "3"]:
            return int(choice)
        print("Lựa chọn không hợp lệ.")


class BinanceRSIAnalyzer:
    def __init__(self):
        load_dotenv()
        self.api_key = os.getenv('BINANCE_API_KEY')
        self.api_secret = os.getenv('BINANCE_API_SECRET')
        self.telegram_token = os.getenv('TELEGRAM_BOT_TOKEN')
        self.telegram_chat_id = os.getenv('TELEGRAM_CHAT_ID')
        self.google_creds_json = os.getenv('GOOGLE_SHEET_CREDENTIALS')
        
        self.client = Client(self.api_key, self.api_secret)
        self.symbols = self._load_symbols()
        self.intervals = ['15m', '1h', '4h', '1d']
        self.rsi_period = 14
        self.excel_file = 'rsi_filtered_data.xlsx'
        self.analysis_mode = 1

        self.RSI_OVERBOUGHT = 80
        self.RSI_OVERSOLD = 20
        self.RSI_UPPER_MID = 60
        self.RSI_LOWER_MID = 40
        self.RSI_CONFIRM_BEARISH = 70
        self.RSI_CONFIRM_BULLISH = 30
        
        self.scan_candles = 100
        self.min_candle_distance = 24
        self.max_candle_distance = 34

    def _load_symbols(self):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(current_dir, 'textcoin.txt')
        with open(file_path, 'r') as file:
            return [line.strip() for line in file.readlines()]

    def _calculate_rsi(self, close_prices, window):
        rsi = ta.momentum.RSIIndicator(close_prices, window=window).rsi()
        return rsi.dropna()

    def _is_local_peak(self, rsi_array, index):
        if index <= 0 or index >= len(rsi_array) - 1:
            return False
        return rsi_array[index] > rsi_array[index - 1] and rsi_array[index] > rsi_array[index + 1]

    def _is_local_bottom(self, rsi_array, index):
        if index <= 0 or index >= len(rsi_array) - 1:
            return False
        return rsi_array[index] < rsi_array[index - 1] and rsi_array[index] < rsi_array[index + 1]

    def _detect_bearish_divergence_v4(self, rsi_array, high_array):
        """
        Bearish Divergence: So sánh giá HIGH tại các đỉnh RSI
        - Giá HIGH tăng nhưng RSI giảm = Bearish Divergence
        """
        n = len(rsi_array)
        if n < self.scan_candles:
            return None

        rsi = rsi_array[-self.scan_candles:]
        highs = high_array[-self.scan_candles:]
        n = len(rsi)

        phase = DivergencePhase.IDLE
        peak1_rsi = None
        peak1_high = None
        peak1_index = None
        peak2_rsi = None
        peak2_high = None
        peak2_index = None
        divergence_ready = False
        
        temp_peak_rsi = None
        temp_peak_high = None
        temp_peak_index = None
        in_overbought = False

        for i in range(n):
            rsi_val = rsi[i]
            high_val = highs[i]

            if phase == DivergencePhase.IDLE:
                if rsi_val > self.RSI_OVERBOUGHT:
                    in_overbought = True
                    if temp_peak_rsi is None or rsi_val > temp_peak_rsi:
                        temp_peak_rsi = rsi_val
                        temp_peak_high = high_val
                        temp_peak_index = i
                elif in_overbought and rsi_val <= self.RSI_OVERBOUGHT:
                    peak1_rsi = temp_peak_rsi
                    peak1_high = temp_peak_high
                    peak1_index = temp_peak_index
                    phase = DivergencePhase.FORMING
                    in_overbought = False
                    temp_peak_rsi = None
                    temp_peak_high = None
                    temp_peak_index = None

            elif phase == DivergencePhase.FORMING:
                if rsi_val < self.RSI_LOWER_MID:
                    phase = DivergencePhase.IDLE
                    peak1_rsi = None
                    peak1_high = None
                    peak1_index = None
                elif rsi_val > self.RSI_OVERBOUGHT:
                    in_overbought = True
                    temp_peak_rsi = rsi_val
                    temp_peak_high = high_val
                    temp_peak_index = i
                    phase = DivergencePhase.IDLE
                    peak1_rsi = None
                    peak1_high = None
                    peak1_index = None
                elif rsi_val < self.RSI_UPPER_MID:
                    phase = DivergencePhase.DEVELOPING

            elif phase == DivergencePhase.DEVELOPING:
                candle_distance = i - peak1_index
                
                if rsi_val < self.RSI_LOWER_MID:
                    phase = DivergencePhase.IDLE
                    peak1_rsi = None
                    peak1_high = None
                    peak1_index = None
                    divergence_ready = False
                elif rsi_val > self.RSI_OVERBOUGHT:
                    in_overbought = True
                    temp_peak_rsi = rsi_val
                    temp_peak_high = high_val
                    temp_peak_index = i
                    phase = DivergencePhase.IDLE
                    peak1_rsi = None
                    peak1_high = None
                    peak1_index = None
                    divergence_ready = False
                elif candle_distance > self.max_candle_distance:
                    phase = DivergencePhase.IDLE
                    peak1_rsi = None
                    peak1_high = None
                    peak1_index = None
                    divergence_ready = False
                elif self.min_candle_distance <= candle_distance <= self.max_candle_distance:
                    if i >= 2 and self._is_local_peak(rsi, i - 1):
                        candidate_index = i - 1
                        candidate_rsi = rsi[candidate_index]
                        candidate_high = highs[candidate_index]
                        
                        if self.RSI_LOWER_MID < candidate_rsi < self.RSI_OVERBOUGHT:
                            # So sánh HIGH: HIGH2 > HIGH1 và RSI2 < RSI1
                            if candidate_high > peak1_high and candidate_rsi < peak1_rsi:
                                peak2_rsi = candidate_rsi
                                peak2_high = candidate_high
                                peak2_index = candidate_index
                                divergence_ready = True
                
                if divergence_ready and rsi_val <= self.RSI_CONFIRM_BEARISH:
                    phase = DivergencePhase.CONFIRMED

        if phase == DivergencePhase.CONFIRMED:
            return {'type': 'bearish', 'stage': 'CONFIRMED'}
        elif phase == DivergencePhase.DEVELOPING and divergence_ready:
            return {'type': 'bearish', 'stage': 'DEVELOPING'}
        elif phase == DivergencePhase.DEVELOPING:
            return {'type': 'bearish', 'stage': 'FORMING'}
        elif phase == DivergencePhase.FORMING:
            return {'type': 'bearish', 'stage': 'FORMING'}

        return None

    def _detect_bullish_divergence_v4(self, rsi_array, low_array):
        """
        Bullish Divergence: So sánh giá LOW tại các đáy RSI
        - Giá LOW giảm nhưng RSI tăng = Bullish Divergence
        """
        n = len(rsi_array)
        if n < self.scan_candles:
            return None

        rsi = rsi_array[-self.scan_candles:]
        lows = low_array[-self.scan_candles:]
        n = len(rsi)

        phase = DivergencePhase.IDLE
        bottom1_rsi = None
        bottom1_low = None
        bottom1_index = None
        bottom2_rsi = None
        bottom2_low = None
        bottom2_index = None
        divergence_ready = False
        
        temp_bottom_rsi = None
        temp_bottom_low = None
        temp_bottom_index = None
        in_oversold = False

        for i in range(n):
            rsi_val = rsi[i]
            low_val = lows[i]

            if phase == DivergencePhase.IDLE:
                if rsi_val < self.RSI_OVERSOLD:
                    in_oversold = True
                    if temp_bottom_rsi is None or rsi_val < temp_bottom_rsi:
                        temp_bottom_rsi = rsi_val
                        temp_bottom_low = low_val
                        temp_bottom_index = i
                elif in_oversold and rsi_val >= self.RSI_OVERSOLD:
                    bottom1_rsi = temp_bottom_rsi
                    bottom1_low = temp_bottom_low
                    bottom1_index = temp_bottom_index
                    phase = DivergencePhase.FORMING
                    in_oversold = False
                    temp_bottom_rsi = None
                    temp_bottom_low = None
                    temp_bottom_index = None

            elif phase == DivergencePhase.FORMING:
                if rsi_val > self.RSI_UPPER_MID:
                    phase = DivergencePhase.IDLE
                    bottom1_rsi = None
                    bottom1_low = None
                    bottom1_index = None
                elif rsi_val < self.RSI_OVERSOLD:
                    in_oversold = True
                    temp_bottom_rsi = rsi_val
                    temp_bottom_low = low_val
                    temp_bottom_index = i
                    phase = DivergencePhase.IDLE
                    bottom1_rsi = None
                    bottom1_low = None
                    bottom1_index = None
                elif rsi_val > self.RSI_LOWER_MID:
                    phase = DivergencePhase.DEVELOPING

            elif phase == DivergencePhase.DEVELOPING:
                candle_distance = i - bottom1_index
                
                if rsi_val > self.RSI_UPPER_MID:
                    phase = DivergencePhase.IDLE
                    bottom1_rsi = None
                    bottom1_low = None
                    bottom1_index = None
                    divergence_ready = False
                elif rsi_val < self.RSI_OVERSOLD:
                    in_oversold = True
                    temp_bottom_rsi = rsi_val
                    temp_bottom_low = low_val
                    temp_bottom_index = i
                    phase = DivergencePhase.IDLE
                    bottom1_rsi = None
                    bottom1_low = None
                    bottom1_index = None
                    divergence_ready = False
                elif candle_distance > self.max_candle_distance:
                    phase = DivergencePhase.IDLE
                    bottom1_rsi = None
                    bottom1_low = None
                    bottom1_index = None
                    divergence_ready = False
                elif self.min_candle_distance <= candle_distance <= self.max_candle_distance:
                    if i >= 2 and self._is_local_bottom(rsi, i - 1):
                        candidate_index = i - 1
                        candidate_rsi = rsi[candidate_index]
                        candidate_low = lows[candidate_index]
                        
                        if self.RSI_OVERSOLD < candidate_rsi < self.RSI_UPPER_MID:
                            # So sánh LOW: LOW2 < LOW1 và RSI2 > RSI1
                            if candidate_low < bottom1_low and candidate_rsi > bottom1_rsi:
                                bottom2_rsi = candidate_rsi
                                bottom2_low = candidate_low
                                bottom2_index = candidate_index
                                divergence_ready = True
                
                if divergence_ready and rsi_val >= self.RSI_CONFIRM_BULLISH:
                    phase = DivergencePhase.CONFIRMED

        if phase == DivergencePhase.CONFIRMED:
            return {'type': 'bullish', 'stage': 'CONFIRMED'}
        elif phase == DivergencePhase.DEVELOPING and divergence_ready:
            return {'type': 'bullish', 'stage': 'DEVELOPING'}
        elif phase == DivergencePhase.DEVELOPING:
            return {'type': 'bullish', 'stage': 'FORMING'}
        elif phase == DivergencePhase.FORMING:
            return {'type': 'bullish', 'stage': 'FORMING'}

        return None

    def _detect_divergence(self, rsi_series, high_series, low_series):
        """
        Phát hiện phân kỳ sử dụng HIGH cho bearish và LOW cho bullish
        """
        rsi = np.array(rsi_series)
        highs = np.array(high_series)
        lows = np.array(low_series)
        
        if len(rsi) < self.scan_candles:
            return []

        results = []

        # Bearish divergence: dùng HIGH
        bearish = self._detect_bearish_divergence_v4(rsi, highs)
        if bearish:
            results.append(bearish)

        # Bullish divergence: dùng LOW
        bullish = self._detect_bullish_divergence_v4(rsi, lows)
        if bullish:
            results.append(bullish)

        return results

    def _fetch_and_process_data(self, symbol, interval):
        try:
            klines = self.client.get_klines(symbol=symbol, interval=interval, limit=200)
            
            # Lấy các loại giá từ klines
            # klines format: [open_time, open, high, low, close, volume, ...]
            high_prices = pd.Series([float(k[2]) for k in klines])
            low_prices = pd.Series([float(k[3]) for k in klines])
            close_prices = pd.Series([float(k[4]) for k in klines])
            current_price = float(klines[-1][4])

            if len(close_prices) < self.rsi_period + self.scan_candles:
                return None

            rsi = self._calculate_rsi(close_prices, self.rsi_period)
            
            # Align tất cả các price series với RSI
            aligned_length = len(rsi)
            aligned_highs = high_prices.iloc[-aligned_length:].reset_index(drop=True)
            aligned_lows = low_prices.iloc[-aligned_length:].reset_index(drop=True)
            aligned_closes = close_prices.iloc[-aligned_length:].reset_index(drop=True)
            
            rsi_last5 = rsi.tail(5)

            # Gọi _detect_divergence với high và low riêng biệt
            divergences = self._detect_divergence(rsi, aligned_highs, aligned_lows)

            bullish_div = None
            bearish_div = None
            for div in divergences:
                if div['type'] == 'bullish':
                    bullish_div = div
                elif div['type'] == 'bearish':
                    bearish_div = div

            return {
                'symbol': symbol,
                'interval': interval,
                'rsi_last5': rsi_last5,
                'current_rsi': round(rsi.iloc[-1], 2),
                'current_price': current_price,
                'divergence_bullish': bullish_div,
                'divergence_bearish': bearish_div
            }
        except Exception as e:
            return None

    def _process_result(self, results):
        output = {interval: {
            'rsi_high': [], 'rsi_low': [],
            'div_bullish_confirmed': [], 'div_bullish_developing': [], 'div_bullish_forming': [],
            'div_bearish_confirmed': [], 'div_bearish_developing': [], 'div_bearish_forming': []
        } for interval in self.intervals}

        for result in results:
            symbol = result['symbol']
            interval = result['interval']
            rsi_last5 = result['rsi_last5']
            chart_url = f'https://www.tradingview.com/chart/?symbol=BINANCE:{symbol}'

            if self.analysis_mode in [1, 3]:
                if (rsi_last5 >= self.RSI_OVERBOUGHT).any():
                    output[interval]['rsi_high'].append({
                        'Tên': symbol,
                        'Loại': f'RSI ≥ {self.RSI_OVERBOUGHT}',
                        'Chart URL': chart_url
                    })
                elif (rsi_last5 <= self.RSI_OVERSOLD).any():
                    output[interval]['rsi_low'].append({
                        'Tên': symbol,
                        'Loại': f'RSI ≤ {self.RSI_OVERSOLD}',
                        'Chart URL': chart_url
                    })

            if self.analysis_mode in [2, 3]:
                if result['divergence_bullish']:
                    div = result['divergence_bullish']
                    key = f'div_bullish_{div["stage"].lower()}'
                    output[interval][key].append({
                        'Tên': symbol,
                        'Loại': '🟢',
                        'Giai đoạn': div['stage'],
                        'Chart URL': chart_url
                    })

                if result['divergence_bearish']:
                    div = result['divergence_bearish']
                    key = f'div_bearish_{div["stage"].lower()}'
                    output[interval][key].append({
                        'Tên': symbol,
                        'Loại': '🔴',
                        'Giai đoạn': div['stage'],
                        'Chart URL': chart_url
                    })

        return output

    def _send_telegram_message(self, data):
        print(f"\n🔍 Debug Telegram:")
        print(f"   Token: {'✓ Có' if self.telegram_token else '✗ Không có'}")
        print(f"   Chat ID: {'✓ Có' if self.telegram_chat_id else '✗ Không có'}")

        if not self.telegram_token or not self.telegram_chat_id:
            print("   ⚠️ Thiếu TELEGRAM_BOT_TOKEN hoặc TELEGRAM_CHAT_ID trong .env")
            return False

        message_parts = [f"🔔 <b>RSI Alert - Period {self.rsi_period}</b>"]
        mode_text = {1: "RSI Cơ bản", 2: "RSI Divergence V4", 3: "RSI + Divergence V4"}
        message_parts.append(f"📊 Mode: {mode_text.get(self.analysis_mode)}")
        message_parts.append(f"📏 Khoảng cách: {self.min_candle_distance}-{self.max_candle_distance} nến")
        message_parts.append(f"💡 Price: HIGH (Bearish) / LOW (Bullish)\n")

        stage_emoji = {'CONFIRMED': '✅', 'DEVELOPING': '🔄', 'FORMING': '🌱'}
        has_any_data = False

        for interval in self.intervals:
            interval_data = data[interval]

            interval_has_data = any([
                interval_data['rsi_high'],
                interval_data['rsi_low'],
                interval_data['div_bullish_confirmed'],
                interval_data['div_bullish_developing'],
                interval_data['div_bullish_forming'],
                interval_data['div_bearish_confirmed'],
                interval_data['div_bearish_developing'],
                interval_data['div_bearish_forming']
            ])

            if interval_has_data:
                has_any_data = True
                message_parts.append(f"\n⏰ <b>Khung {interval}</b>")

                if self.analysis_mode in [1, 3]:
                    if interval_data['rsi_high']:
                        message_parts.append(f"\n📈 <b>RSI ≥ {self.RSI_OVERBOUGHT}:</b>")
                        for item in interval_data['rsi_high'][:10]:
                            message_parts.append(f"• {item['Tên']} | <a href='{item['Chart URL']}'>Chart</a>")

                    if interval_data['rsi_low']:
                        message_parts.append(f"\n📉 <b>RSI ≤ {self.RSI_OVERSOLD}:</b>")
                        for item in interval_data['rsi_low'][:10]:
                            message_parts.append(f"• {item['Tên']} | <a href='{item['Chart URL']}'>Chart</a>")

                if self.analysis_mode in [2, 3]:
                    for stage in ['confirmed', 'developing', 'forming']:
                        key = f'div_bullish_{stage}'
                        if interval_data[key]:
                            message_parts.append(f"\n🟢 <b>BULLISH DIV - {stage_emoji[stage.upper()]} {stage.upper()}:</b>")
                            for item in interval_data[key][:10]:
                                message_parts.append(f"• {item['Tên']} | <a href='{item['Chart URL']}'>Chart</a>")

                    for stage in ['confirmed', 'developing', 'forming']:
                        key = f'div_bearish_{stage}'
                        if interval_data[key]:
                            message_parts.append(f"\n🔴 <b>BEARISH DIV - {stage_emoji[stage.upper()]} {stage.upper()}:</b>")
                            for item in interval_data[key][:10]:
                                message_parts.append(f"• {item['Tên']} | <a href='{item['Chart URL']}'>Chart</a>")

        if not has_any_data:
            message_parts.append("\n✅ Không có tín hiệu nào")

        message = "\n".join(message_parts)

        if len(message) > 4000:
            message = message[:4000] + "\n\n... (truncated)"

        url = f"https://api.telegram.org/bot{self.telegram_token}/sendMessage"
        payload = {
            "chat_id": self.telegram_chat_id,
            "text": message,
            "parse_mode": "HTML",
            "disable_web_page_preview": True
        }

        try:
            print(f"   📤 Đang gửi message ({len(message)} ký tự)...")
            response = requests.post(url, json=payload, timeout=30)

            if response.status_code == 200:
                print("   ✅ Telegram: Gửi thành công!")
                return True
            else:
                print(f"   ❌ Telegram Error: Status {response.status_code}")
                print(f"   Response: {response.text}")
                return False

        except requests.exceptions.Timeout:
            print("   ❌ Telegram: Timeout")
            return False
        except requests.exceptions.ConnectionError:
            print("   ❌ Telegram: Connection Error")
            return False
        except Exception as e:
            print(f"   ❌ Telegram Exception: {str(e)}")
            return False

    def _save_to_excel(self, data):
        with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
            for interval in self.intervals:
                rows = []

                if self.analysis_mode in [1, 3]:
                    for item in data[interval]['rsi_high']:
                        rows.append({
                            'Tên': item['Tên'],
                            'Loại': item['Loại'],
                            'Giai đoạn': '-',
                            'Chart URL': item['Chart URL']
                        })
                    for item in data[interval]['rsi_low']:
                        rows.append({
                            'Tên': item['Tên'],
                            'Loại': item['Loại'],
                            'Giai đoạn': '-',
                            'Chart URL': item['Chart URL']
                        })

                if self.analysis_mode in [2, 3]:
                    for stage in ['confirmed', 'developing', 'forming']:
                        for item in data[interval][f'div_bullish_{stage}']:
                            rows.append({
                                'Tên': item['Tên'],
                                'Loại': item['Loại'],
                                'Giai đoạn': item['Giai đoạn'],
                                'Chart URL': item['Chart URL']
                            })
                        for item in data[interval][f'div_bearish_{stage}']:
                            rows.append({
                                'Tên': item['Tên'],
                                'Loại': item['Loại'],
                                'Giai đoạn': item['Giai đoạn'],
                                'Chart URL': item['Chart URL']
                            })

                columns = ['Tên', 'Loại', 'Giai đoạn', 'Chart URL']
                df = pd.DataFrame(rows, columns=columns) if rows else pd.DataFrame(columns=columns)
                df.to_excel(writer, sheet_name=interval, index=False)

                ws = writer.sheets[interval]
                for col in ws.columns:
                    max_length = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min((max_length + 2) * 1.2, 80)

                border_style = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = border_style

                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')

                stage_fills = {
                    'CONFIRMED': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                    'DEVELOPING': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
                    'FORMING': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                }

                header_row = [cell.value for cell in ws[1]]
                stage_col = header_row.index('Giai đoạn') + 1 if 'Giai đoạn' in header_row else None

                for row_idx in range(2, ws.max_row + 1):
                    if stage_col:
                        stage_cell = ws.cell(row=row_idx, column=stage_col)
                        if stage_cell.value in stage_fills:
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).fill = stage_fills[stage_cell.value]

                ws.auto_filter.ref = ws.dimensions
                ws.freeze_panes = 'A2'

        print(f'✅ Excel saved: {self.excel_file}')

    def _upload_to_google_sheet(self, data):
        try:
            scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
            creds = ServiceAccountCredentials.from_json_keyfile_name(self.google_creds_json, scope)
            client_sheet = gspread.authorize(creds)
            spreadsheet = client_sheet.open("RSI Data")

            for interval in self.intervals:
                try:
                    worksheet = spreadsheet.worksheet(interval)
                except gspread.exceptions.WorksheetNotFound:
                    worksheet = spreadsheet.add_worksheet(title=interval, rows="100", cols="20")

                headers = ["Tên", "Loại", "Giai đoạn", "Chart URL"]
                values = [headers]

                if self.analysis_mode in [1, 3]:
                    for item in data[interval].get('rsi_high', []):
                        values.append([item['Tên'], item['Loại'], '-', item['Chart URL']])
                    for item in data[interval].get('rsi_low', []):
                        values.append([item['Tên'], item['Loại'], '-', item['Chart URL']])

                if self.analysis_mode in [2, 3]:
                    for stage in ['confirmed', 'developing', 'forming']:
                        for item in data[interval].get(f'div_bullish_{stage}', []):
                            values.append([
                                item['Tên'], item['Loại'], item['Giai đoạn'], item['Chart URL']
                            ])
                        for item in data[interval].get(f'div_bearish_{stage}', []):
                            values.append([
                                item['Tên'], item['Loại'], item['Giai đoạn'], item['Chart URL']
                            ])

                worksheet.clear()
                worksheet.update("A1", values)

            print("✅ Google Sheet updated!")
        except FileNotFoundError:
            print("❌ Không tìm thấy file your.json")
        except Exception as e:
            print(f"❌ Google Sheet error: {str(e)}")

    def analyze(self):
        mode = ask_analysis_mode()
        if mode is None:
            return
        self.analysis_mode = mode

        while True:
            period = ask_rsi_period()
            if period is not None:
                self.rsi_period = period
                break

        while True:
            intervals = ask_intervals()
            if intervals is not None:
                self.intervals = intervals
                break

        print(f"\n{'='*60}")
        print(f"📊 Mode: {self.analysis_mode} | RSI: {self.rsi_period} | Intervals: {self.intervals}")
        print(f"📏 Khoảng cách phân kỳ: {self.min_candle_distance}-{self.max_candle_distance} nến")
        print(f"🔍 Scan: {self.scan_candles} nến gần nhất")
        print(f"💡 Price comparison: HIGH (Bearish) / LOW (Bullish)")
        print(f"⚙️ RSI Thresholds: OB={self.RSI_OVERBOUGHT} | OS={self.RSI_OVERSOLD} | Mid={self.RSI_LOWER_MID}-{self.RSI_UPPER_MID}")
        print(f"⚙️ Confirm: Bearish={self.RSI_CONFIRM_BEARISH} | Bullish={self.RSI_CONFIRM_BULLISH}")
        print(f"{'='*60}\n")

        results = []
        for interval in self.intervals:
            print(f'🔄 Processing {interval}...')
            with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
                futures = {
                    executor.submit(self._fetch_and_process_data, symbol, interval): symbol
                    for symbol in self.symbols
                }
                completed = 0
                total = len(self.symbols)
                for future in concurrent.futures.as_completed(futures):
                    completed += 1
                    print(f'\r📊 {interval}: {(completed/total)*100:.1f}%', end='', flush=True)
                    result = future.result()
                    if result:
                        results.append(result)
            print(f'\n✅ Done {interval}!')

        processed_data = self._process_result(results)

        print(f"\n{'='*70}")
        print("📊 SUMMARY:")
        for interval in self.intervals:
            print(f"\n⏰ {interval}:")
            if self.analysis_mode in [1, 3]:
                print(f"   📈 RSI ≥ {self.RSI_OVERBOUGHT}: {len(processed_data[interval]['rsi_high'])}")
                print(f"   📉 RSI ≤ {self.RSI_OVERSOLD}: {len(processed_data[interval]['rsi_low'])}")
            if self.analysis_mode in [2, 3]:
                bull_c = len(processed_data[interval]['div_bullish_confirmed'])
                bull_d = len(processed_data[interval]['div_bullish_developing'])
                bull_f = len(processed_data[interval]['div_bullish_forming'])
                bear_c = len(processed_data[interval]['div_bearish_confirmed'])
                bear_d = len(processed_data[interval]['div_bearish_developing'])
                bear_f = len(processed_data[interval]['div_bearish_forming'])
                print(f"   🟢 Bullish Div: C={bull_c} D={bull_d} F={bull_f}")
                print(f"   🔴 Bearish Div: C={bear_c} D={bear_d} F={bear_f}")
        print(f"{'='*70}\n")

        self._save_to_excel(processed_data)
        self._upload_to_google_sheet(processed_data)
        self._send_telegram_message(processed_data)

        print(f'\n🔥 Complete!')


if __name__ == "__main__":
    analyzer = BinanceRSIAnalyzer()
    analyzer.analyze()
