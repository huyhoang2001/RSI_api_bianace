import os
import concurrent.futures
from binance.client import Client
import pandas as pd
import numpy as np
from dotenv import load_dotenv
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from colorama import Fore, Style, init
import ta

init(autoreset=True)
load_dotenv()

class BinanceRSIAnalyzer:
    def __init__(self):
        self.api_key = os.getenv('BINANCE_API_KEY')
        self.api_secret = os.getenv('BINANCE_API_SECRET')
        self.client = Client(self.api_key, self.api_secret)
        self.symbols = self._load_symbols()
        self.intervals = ['1h', '4h','1d']
        self.excel_file = 'rsi_filtered_data.xlsx'

    def _load_symbols(self):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(current_dir, 'textcoin.txt')
        with open(file_path, 'r') as file:
            return [line.strip() for line in file.readlines()]

    def _calculate_rsi(self, close_prices, window=7):
        rsi = ta.momentum.RSIIndicator(close_prices, window=window).rsi()
        return rsi.dropna().tail(5)

    def _fetch_and_process_data(self, symbol, interval):
        try:
            klines = self.client.get_klines(
                symbol=symbol,
                interval=interval,
                limit=500
            )
            
            close_prices = pd.Series([float(k[4]) for k in klines])
            if len(close_prices) < 7:
                return None

            rsi = self._calculate_rsi(close_prices)
            return (symbol, interval, rsi)
        except Exception as e:
            print(f"Error processing {symbol} {interval}: {str(e)}")
            return None

    def _process_result(self, results):
        output = {interval: {'rsi_high': [], 'rsi_low': []} for interval in self.intervals}
        
        for symbol, interval, rsi in results:
            if (rsi >= 80).any():
                output[interval]['rsi_high'].append({
                    'Tên': symbol,
                    'Chart URL': f'https://www.tradingview.com/chart/?symbol=BINANCE:{symbol}'
                })
            elif (rsi <= 20).any():
                output[interval]['rsi_low'].append({
                    'Tên': symbol,
                    'Chart URL': f'https://www.tradingview.com/chart/?symbol=BINANCE:{symbol}'
                })
        return output

    def _save_to_excel(self, data):   
        with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
            for interval in self.intervals:
                # Tạo DataFrame kết hợp
                high_df = pd.DataFrame(data[interval]['rsi_high'])
                low_df = pd.DataFrame(data[interval]['rsi_low'])
                
                # Tạo spacer với đầy đủ cột
                spacer = pd.DataFrame([[''] * 3], columns=['Tên', 'Chart URL', 'RSI Condition'])
                
                combined_df = pd.concat([
                    high_df.assign(**{'RSI Condition': '>=80'}),
                    spacer,
                    low_df.assign(**{'RSI Condition': '<=20'})
                ], ignore_index=True)

                # Ghi vào Excel
                combined_df.to_excel(
                    writer,
                    sheet_name=interval,
                    index=False,
                    startrow=0
                )

                # Lấy worksheet và áp dụng định dạng
                ws = writer.sheets[interval]
                
                # Tự động điều chỉnh cột
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column].width = adjusted_width

                # Tạo kiểu border đậm
                border_style = Border(left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin'))

                # Áp dụng border cho toàn bộ dữ liệu
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = border_style

                # In đậm header và canh giữa
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')

                # Thêm filter và đóng băng hàng tiêu đề
                ws.auto_filter.ref = ws.dimensions
                ws.freeze_panes = 'A2'

        print(f'File Excel đã được cập nhật với định dạng chuyên nghiệp!')

    def analyze(self):
        processed_data = {interval: {'rsi_high': [], 'rsi_low': []} for interval in self.intervals}
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            for interval in self.intervals:
                print(f'🔄 Đang xử lý khung thời gian {interval}...')
                
                # Tạo danh sách công việc cho interval hiện tại
                futures = [executor.submit(self._fetch_and_process_data, symbol, interval) 
                        for symbol in self.symbols]
                
                # Hiển thị thanh tiến trình
                completed = 0
                total = len(self.symbols)
                
                while len(futures) > 0:
                    # Chờ các task hoàn thành
                    done, futures = concurrent.futures.wait(
                        futures, 
                        return_when=concurrent.futures.FIRST_COMPLETED
                    )
                    
                    # Cập nhật tiến trình
                    completed += len(done)
                    progress = (completed / total) * 100
                    print(f'\r📊 Tiến trình {interval}: {progress:.1f}%', end='', flush=True)
                    
                    # Xử lý kết quả
                    for future in done:
                        result = future.result()
                        if result:
                            symbol, interval_used, rsi_values = result
                            # Thêm vào dữ liệu đã xử lý
                            if (rsi_values >= 80).any():
                                processed_data[interval_used]['rsi_high'].append({
                                    'Tên': symbol,
                                    'Chart URL': f'https://www.tradingview.com/chart/?symbol=BINANCE:{symbol}'
                                })
                            elif (rsi_values <= 20).any():
                                processed_data[interval_used]['rsi_low'].append({
                                    'Tên': symbol,
                                    'Chart URL': f'https://www.tradingview.com/chart/?symbol=BINANCE:{symbol}'
                                })
                
                print(f'\n✅ Đã hoàn thành khung {interval}!')
        
        self._save_to_excel(processed_data)
        print(f'🔥 Tất cả dữ liệu đã được lưu vào {self.excel_file}')

if __name__ == "__main__":
    analyzer = BinanceRSIAnalyzer()
    analyzer.analyze()
